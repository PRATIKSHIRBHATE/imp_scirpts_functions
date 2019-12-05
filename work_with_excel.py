# -*- coding: utf-8 -*-
"""
Created on Fri Feb 15 18:50:12 2019

@author: PShirbhate
"""

import openpyxl

sheet_names = ['A','B','C']

wb = openpyxl.load_workbook('data.xlsx')

for x in sheet_names:
    wb.remove_sheet(wb.get_sheet_by_name(x))
    wb.create_sheet(title = x)

wb.save('data1.xlsx')
wb.close()


import openpyxl

sheet_names = ['A','B','C']

wb = openpyxl.load_workbook('data.xlsx')

for x in sheet_names:
    wb.remove(wb[x])
    wb.create_sheet(title = x)

wb.save('data1.xlsx')
wb.close()




def multply(a,b):
    try:
        c = a*b
    except ValueError as err:
        print("can't multiply sequence by non-int of type 'str")
    return c

multply('d','e')

